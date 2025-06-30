import jsonlines
import json
import shutil
import fitz
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
import camelot
from itertools import permutations
import os

#生成pdf和excel文件到all文件夹
def create_pdf_and_excel():

    #获取单元格逻辑始末位置
    def found_cell_loc(cell_id, cell_ID_matrix):
        loc = []
        for i in range(len(cell_ID_matrix)):
            for j in range(len(cell_ID_matrix[0])):
                if cell_ID_matrix[i][j] == cell_id:
                    loc.append([i, j])
        max_x = 0
        max_y = 0
        min_x = 10000
        min_y = 10000
        for i in range(len(loc)):
            if loc[i][0] > max_x:
                max_x = loc[i][0]
            if loc[i][1] > max_y:
                max_y = loc[i][1]
            if loc[i][0] < min_x:
                min_x = loc[i][0]
            if loc[i][1] < min_y:
                min_y = loc[i][1]
        return min_x + 1, min_y + 1, max_x + 1, max_y + 1

    #获取列索引，用于生成excel
    def get_index(i):
        index = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        if i <= 25:
            result = index[i]
        else:
            result = index[int(i / 26) - 1] + index[i % 26]

        return result

    #生成excel文件
    def create_excel(table):
        workbook = Workbook()
        booksheet = workbook.active
        cell_num = len(table['chinese_cell_value_list'])
        for cell_id in range(cell_num):
            start_row, start_column, end_row, end_column = found_cell_loc(cell_id, table['cell_ID_matrix'])
            booksheet.cell(start_row, start_column).value = table['chinese_cell_value_list'][cell_id]
            booksheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
        # 试出的最大值
        max_len = 87
        ave_len = max_len / len(table['cell_ID_matrix'][0])
        for i in range(len(table['cell_ID_matrix'][0])):
            booksheet.column_dimensions[get_index(i)].width = ave_len
        # 遍历所有行和列，设置单元格自动换行和边框
        for row in booksheet.iter_rows():
            for cell in row:
                # 自动换行，上下左右归中
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        workbook.save('./imtsrtqa-noline/all/' + table['table_id'] + '.xlsx')

        # 遍历所有行和列，设置单元格自动换行和边框
        for row in booksheet.iter_rows():
            for cell in row:
                # 自动换行，上下左右归中
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # 边框
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
        workbook.save('./imtsrtqa-line/all/' + table['table_id'] + '.xlsx')

    def convert_to_pdf(xlsx_file, output_dir):
        libreoffice_path = r"C:\Program Files\LibreOffice\program\soffice.exe"
        cmd = [libreoffice_path, '--headless', '--convert-to', 'pdf', '--outdir', output_dir, output_dir + xlsx_file]
        subprocess.run(cmd, check=True, capture_output=True)

    with open('./data/train_tables.json', 'r', encoding='utf-8') as file:
        train_tables = json.load(file)
    with open('./data/test_tables.json', 'r', encoding='utf-8') as file:
        test_tables = json.load(file)
    tables = train_tables + test_tables
    for i in range(len(tables)):
        create_excel(tables[i])
        convert_to_pdf(tables[i]['table_id'] + '.xlsx', './imtsrtqa-line/all/')
        convert_to_pdf(tables[i]['table_id'] + '.xlsx', './imtsrtqa-noline/all/')

#获取数据集标签
def get_lables():

    def pdf_to_images(pdf_path, output_folder, zoom=3):
        doc = fitz.open(pdf_path)
        for page_num, page in enumerate(doc):
            # 设置缩放因子（提高分辨率）
            matrix = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            img_path = f"{output_folder}_{page_num + 1}.png"
            pix.save(img_path)
        doc.close()

    #作为ocr识别结果
    def get_vision_truth(file_name):
        try:
            camelot.read_pdf('./imtsrtqa-noline/all/'+file_name, flavor="stream", pages='2')
            page_num=2
        except:
            page_num=1
        if page_num==2:
            return 0,0
        tables = camelot.read_pdf('./imtsrtqa-noline/all/'+file_name, flavor="stream", pages='1')
        # 检查是否找到表格
        if tables.n == 0:
            return 0,0
        else:
            cells_list=[]
            text_list=[]
            for table_idx, table in enumerate(tables):
                cells = table.cells
                for row_idx in range(table.df.shape[0]):
                    for col_idx in range(table.df.shape[1]):
                        # 获取单元格文本
                        cell_text = table.df.iloc[row_idx, col_idx].strip().replace('\n','').replace('\t','').replace(' ','').replace('\xa0','')
                        # 获取单元格坐标（四个顶点坐标）
                        cell_coords = cells[row_idx][col_idx]
                        x1, y1 = cell_coords.lt  # 左上角
                        x2, y2 = cell_coords.rt  # 右上角
                        x3, y3 = cell_coords.rb  # 右下角
                        x4, y4 = cell_coords.lb  # 左下角
                        if cell_text!='':
                            cells_list.append([x1,y1,x2,y2,x3,y3,x4,y4])
                            text_list.append(cell_text)
            return cells_list,text_list

    #作为真实标签结果
    def get_ground_truth(file_name):
        try:
            camelot.read_pdf('./imtsrtqa-line/all/'+file_name, flavor="lattice", pages='2')
            page_num=2
        except:
            page_num=1
        if page_num==2:
            return 0,0
        tables = camelot.read_pdf('./imtsrtqa-line/all/'+file_name, flavor="lattice", pages='1')
        # 检查是否找到表格
        if tables.n == 0:
            return 0,0
        else:
            cells_list=[]
            text_list=[]
            for table_idx, table in enumerate(tables):
                cells = table.cells
                for row_idx in range(table.df.shape[0]):
                    for col_idx in range(table.df.shape[1]):
                        # 获取单元格文本
                        cell_text = table.df.iloc[row_idx, col_idx].strip().replace('\n','').replace('\t','').replace(' ','').replace('\xa0','')
                        # 获取单元格坐标（四个顶点坐标）
                        cell_coords = cells[row_idx][col_idx]
                        x1, y1 = cell_coords.lt  # 左上角
                        x2, y2 = cell_coords.rt  # 右上角
                        x3, y3 = cell_coords.rb  # 右下角
                        x4, y4 = cell_coords.lb  # 左下角
                        if cell_text!='':
                            cells_list.append([x1, y1, x2, y2, x3, y3, x4, y4])
                            text_list.append(cell_text)
            return cells_list,text_list

    #判断字符个数是否相同
    def is_char_count_same(str1, str2):
        # 统计第一个字符串中每个字符出现的次数
        char_count1 = {}
        for char in str1:
            if char in char_count1:
                char_count1[char] += 1
            else:
                char_count1[char] = 1

        # 统计第二个字符串中每个字符出现的次数
        char_count2 = {}
        for char in str2:
            if char in char_count2:
                char_count2[char] += 1
            else:
                char_count2[char] = 1

        # 比较两个字典是否相等
        return char_count1 == char_count2

    # 输入一个数组长度，返回所有可能的排列组合
    def generate_all_combinations(length):
        all_combinations = []
        for r in range(1, length + 1):
            # 生成所有可能的排列
            perms = permutations(range(length), r)
            all_combinations = all_combinations + [list(perm) for perm in perms]
        return all_combinations

    #按照icdar数据集的样式生成数据集
    with open('./data/train_tables.json', 'r', encoding='utf-8') as file:
        train_tables = json.load(file)
    with open('./data/test_tables.json', 'r', encoding='utf-8') as file:
        test_tables = json.load(file)
    tables = train_tables + test_tables
    json_file=[]
    vision_json=[]
    combine_list=''
    tsr_list=''
    for i in range(len(tables)):
        table=tables[i]
        ground_cells, ground_texts=get_ground_truth(table['table_id']+'.pdf')
        vision_cells, vision_texts=get_vision_truth(table['table_id']+'.pdf')

        if ground_texts==0:
            continue

        loc_lg=[]#单元格逻辑位置
        chinese_table=[]#真实单元格文本
        for j in range(len(table['chinese_cell_value_list'])):
            temp=table['chinese_cell_value_list'][j].replace(' ','').replace('\n','').replace('\t','').replace('\xa0','')
            if temp!='':
                chinese_table.append(temp)
                sc,sr,ec,er=1000,1000,0,0
                for k in range(len(table['cell_ID_matrix'])):
                    for l in range(len(table['cell_ID_matrix'][k])):
                        if table['cell_ID_matrix'][k][l]==j:
                            if k<sr:
                                sr=k
                            if k>er:
                                er=k
                            if l<sc:
                                sc=l
                            if l>ec:
                                ec=l
                loc_lg.append([sr,sc,er,ec])
        #获取既有真实标签，又有ocr识别标签的表格，用于合并实验
        vision_flag=False
        ground_flag=False

        #生成数据集到xml和image和pdf，有线的表格
        if chinese_table==ground_texts:

            ground_flag=True
            # 生成图像
            pdf_to_images('./imtsrtqa-line/all/'+table['table_id']+'.pdf','./imtsrtqa-line/image/'+table['table_id'])
            # 复制pdf文件
            shutil.copy('./imtsrtqa-line/all/'+table['table_id']+'.pdf', './imtsrtqa-line/pdf/'+table['table_id']+'.pdf')
            # 生成xml文件
            xml_string = '<document><table><region page="1">'
            for k in range(len(chinese_table)):
                sc,sr,ec,er=loc_lg[k][1],loc_lg[k][0],loc_lg[k][3],loc_lg[k][2]
                x1,x2,y1,y2=int(ground_cells[k][0]),int(ground_cells[k][2]),int(ground_cells[k][1]),int(ground_cells[k][5])
                tt=ground_texts[k].replace('<', '&lt;').replace('>', '&gt;').replace('&', '&amp;').replace('"', '&quot;').replace("'", '&apos;')
                xml_string = xml_string + f'<cell id="1" start-col="{sc}" start-row="{sr}" end-col="{ec}" end-row="{er}">'
                xml_string = xml_string + f'<bounding-box x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}"/>'
                xml_string = xml_string + f'<content>{tt}</content>'
                xml_string = xml_string + '</cell>'
            xml_string = xml_string + '</region></table></document>'
            file_path = './imtsrtqa-line/xml/'+table['table_id']+'-str.xml'
            with open(file_path, 'w',encoding='utf-8') as file:
                file.write(xml_string)



        #生成数据集到xml和image和pdf，无线的表格，作为ocr识别结果
        # 生成图像
        pdf_to_images('./imtsrtqa-noline/all/'+table['table_id']+'.pdf','./imtsrtqa-noline/image/'+table['table_id'])
        # 复制pdf文件
        shutil.copy('./imtsrtqa-noline/all/'+table['table_id']+'.pdf', './imtsrtqa-noline/pdf/'+table['table_id']+'.pdf')
        if is_char_count_same(''.join(vision_texts), ''.join(ground_texts)):
            vision_flag=True
            # 生成xml文件
            xml_string = '<document><table><region page="1">'
            for k in range(len(vision_cells)):
                x1,x2,y1,y2=int(vision_cells[k][0]),int(vision_cells[k][2]),int(vision_cells[k][1]),int(vision_cells[k][5])
                tt=vision_texts[k].replace('<', '&lt;').replace('>', '&gt;').replace('&', '&amp;').replace('"', '&quot;').replace("'", '&apos;')
                xml_string = xml_string + f'<cell id="1" start-col="{0}" start-row="{0}">'
                xml_string = xml_string + f'<bounding-box x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}"/>'
                xml_string = xml_string + f'<content>{tt}</content>'
                xml_string = xml_string + '</cell>'
            xml_string = xml_string + '</region></table></document>'
            file_path = './imtsrtqa-noline/xml/'+table['table_id']+'-str.xml'
            with open(file_path, 'w',encoding='utf-8') as file:
                file.write(xml_string)

        #生成结构识别数据集
        if ground_flag:
            tsr_list=tsr_list+ table['table_id']+'_1_1_0'+'\n'
            # 复制pdf文件
            shutil.copy('./imtsrtqa-noline/pdf/' + table['table_id'] + '.pdf','./imtsrtqa-tsr-test/eu-us-dataset/' + table['table_id'] + '.pdf')
            # 复制xml文件
            shutil.copy('./imtsrtqa-line/xml/' + table['table_id'] + '-str.xml','./imtsrtqa-tsr-test/eu-us-dataset/' + table['table_id'] + '-str.xml')
            # 复制image文件
            shutil.copy('./imtsrtqa-noline/image/' + table['table_id'] + '_1.png','./imtsrtqa-tsr-test/src_page_image/' + table['table_id'] + '_1.png')


        #生成合并数据集
        combine_flag=True
        if not ground_flag or not vision_flag:
            combine_flag=False
        for j in range(len(vision_texts)):
            if any(vision_texts[j] in s for s in ground_texts):
                pass
            else:
                combine_flag=False

        if combine_flag:
            vision_json.append({'table_id':table['table_id'],'text':vision_texts})
            combine_list=combine_list + table['table_id']+'_1_1_0'+'\n'
            # 复制pdf文件
            shutil.copy('./imtsrtqa-noline/pdf/' + table['table_id'] + '.pdf','./imtsrtqa-ocr-combine/eu-us-dataset/' + table['table_id'] + '.pdf')
            # 复制xml文件
            shutil.copy('./imtsrtqa-noline/xml/' + table['table_id'] + '-str.xml','./imtsrtqa-ocr-combine/eu-us-dataset/' + table['table_id'] + '-str.xml')
            # 复制image文件
            shutil.copy('./imtsrtqa-noline/image/' + table['table_id'] + '_1.png','./imtsrtqa-ocr-combine/src_page_image/' + table['table_id'] + '_1.png')
            for j in range(len(ground_texts)):
                if ground_texts[j] not in vision_texts:

                    # 如果单元格不存在，找到最近的组合成单元格的文本块

                    temp_text = []
                    temp_bbox = []
                    temp_index = []
                    for k in range(len(vision_texts)):
                        if vision_texts[k] in ground_texts[j]:
                            if vision_cells[k][2] > ground_cells[j][0] and vision_cells[k][0] < ground_cells[j][
                                2] and vision_cells[k][1] > ground_cells[j][5] and vision_cells[k][5] < \
                                    ground_cells[j][1]:
                                temp_text.append(vision_texts[k])
                                temp_bbox.append(vision_cells[k])
                                temp_index.append(k)

                    # 一个都没有或者大于10不要
                    if len(temp_index) == 0:
                        continue
                    if len(temp_index) >= 10:
                        continue

                    all_combinations = generate_all_combinations(len(temp_index))
                    # 得到所有正确组合
                    right_combinations = []
                    for idd, combinations in enumerate(all_combinations):
                        # 得到一个排列可以组成的文本
                        tt = ''
                        for k in range(len(combinations)):
                            tt = tt + temp_text[combinations[k]]
                        # 判断是否组成正确
                        if tt == ground_texts[j]:
                            right_combinations.append(combinations)

                    # 没有正确组合
                    if len(right_combinations) == 0:
                        continue

                    # 计算所有正确组合的距离
                    right_combinations_distance = [0] * len(right_combinations)
                    for idd, combinations in enumerate(right_combinations):
                        for k in range(len(combinations)):
                            dd = (abs((temp_bbox[combinations[k]][0] + temp_bbox[combinations[k]][2]) / 2 -
                                      (ground_cells[j][0] + ground_cells[j][2]) / 2) +
                                  abs((temp_bbox[combinations[k]][1] + temp_bbox[combinations[k]][5]) / 2 -
                                      (ground_cells[j][1] + ground_cells[j][5]) / 2))
                            right_combinations_distance[idd] = right_combinations_distance[idd] + dd
                    # 计算最小距离的索引
                    min_distance_index = 0
                    for k in range(len(right_combinations_distance)):
                        if right_combinations_distance[k] < right_combinations_distance[min_distance_index]:
                            min_distance_index = k
                    right_combination = right_combinations[min_distance_index]
                    right_index = []
                    for i in range(len(right_combination)):
                        right_index.append(temp_index[right_combination[i]])

                    json_file.append({'table_id': table['table_id'], 'combine': right_index})

    with jsonlines.open('./imtsrtqa-ocr-combine/combine.jsonl', mode='w') as writer:
        writer.write_all(json_file)
    #前20%作为测试集
    with open('./imtsrtqa-ocr-combine/src_set/test.txt', 'w', encoding='utf-8') as f:
        f.write(combine_list[0:int(len(combine_list)/5)])

    with open('./imtsrtqa-tsr-test/src_set/test.txt', 'w', encoding='utf-8') as f:
        f.write(tsr_list[0:int(len(tsr_list)/5)])

    with jsonlines.open('./imtsrtqa-ocr-combine/vision.jsonl', mode='w') as writer:
        writer.write_all(vision_json)

if __name__ == '__main__':


    if not os.path.exists('./imtsrtqa-noline/'):
        os.mkdir('./imtsrtqa-noline/')
    if not os.path.exists('./imtsrtqa-noline/all/'):
        os.mkdir('./imtsrtqa-noline/all/')
    if not os.path.exists('./imtsrtqa-noline/image/'):
        os.mkdir('./imtsrtqa-noline/image/')
    if not os.path.exists('./imtsrtqa-noline/pdf/'):
        os.mkdir('./imtsrtqa-noline/pdf/')
    if not os.path.exists('./imtsrtqa-noline/xml/'):
        os.mkdir('./imtsrtqa-noline/xml/')

    if not os.path.exists('./imtsrtqa-line/'):
        os.mkdir('./imtsrtqa-line/')
    if not os.path.exists('./imtsrtqa-line/all/'):
        os.mkdir('./imtsrtqa-line/all/')
    if not os.path.exists('./imtsrtqa-line/image/'):
        os.mkdir('./imtsrtqa-line/image/')
    if not os.path.exists('./imtsrtqa-line/pdf/'):
        os.mkdir('./imtsrtqa-line/pdf/')
    if not os.path.exists('./imtsrtqa-line/xml/'):
        os.mkdir('./imtsrtqa-line/xml/')

    if not os.path.exists('./imtsrtqa-ocr-combine/'):
        os.mkdir('./imtsrtqa-ocr-combine/')
    if not os.path.exists('./imtsrtqa-ocr-combine/eu-us-dataset/'):
        os.mkdir('./imtsrtqa-ocr-combine/eu-us-dataset/')
    if not os.path.exists('./imtsrtqa-ocr-combine/src_set/'):
        os.mkdir('./imtsrtqa-ocr-combine/src_set/')
    if not os.path.exists('./imtsrtqa-ocr-combine/src_page_image/'):
        os.mkdir('./imtsrtqa-ocr-combine/src_page_image/')

    if not os.path.exists('./imtsrtqa-tsr-test/'):
        os.mkdir('./imtsrtqa-tsr-test/')
    if not os.path.exists('./imtsrtqa-tsr-test/eu-us-dataset/'):
        os.mkdir('./imtsrtqa-tsr-test/eu-us-dataset/')
    if not os.path.exists('./imtsrtqa-tsr-test/src_set/'):
        os.mkdir('./imtsrtqa-tsr-test/src_set/')
    if not os.path.exists('./imtsrtqa-tsr-test/src_page_image/'):
        os.mkdir('./imtsrtqa-tsr-test/src_page_image/')

    #create_pdf_and_excel()
    get_lables()